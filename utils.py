from datetime import datetime, timedelta
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def generate_excel_report(db, period):
    now = datetime.now()
    
    if period == 'day':
        start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
        period_name = "за сегодня"
    elif period == 'week':
        start_date = now - timedelta(days=7)
        period_name = "за неделю"
    elif period == 'month':
        start_date = now - timedelta(days=30)
        period_name = "за месяц"
    elif period == 'year':
        start_date = now - timedelta(days=365)
        period_name = "за год"
    
    cursor = db.conn.cursor()
    cursor.execute('''
    SELECT 
        last_name, first_name, patronymic, birth_date, phone_number,
        military_spec, dental_sanation, medical_certificates,
        foreign_passport, active_contracts, registration_date
    FROM users 
    WHERE registration_date >= ? 
    ORDER BY registration_date DESC
    ''', (start_date,))
    
    data = cursor.fetchall()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет"
    
    # Заголовок отчета
    ws['A1'] = f"Отчет по регистрациям {period_name}"
    ws.merge_cells('A1:K1')
    title_cell = ws['A1']
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Заголовки столбцов
    headers = [
        'Фамилия', 'Имя', 'Отчество', 'Дата рождения', 'Телефон',
        'ВУС и профессия', 'Санация', 'Справки', 'Загранпаспорт',
        'Контракты', 'Дата регистрации'
    ]
    
    # Стили для заголовков и ячеек
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    header_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Применяем заголовки и стили
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
    
    # Записываем данные
    for row_idx, row in enumerate(data, 3):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            # Преобразуем булевы значения
            if isinstance(value, int) and col_idx in [7, 8, 9, 10]:
                cell.value = "Да" if value else "Нет"
            else:
                cell.value = value
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border
    
    # Устанавливаем ширину столбцов
    column_widths = [15, 15, 15, 15, 15, 30, 10, 10, 12, 10, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    filename = f'report_{period}_{now.strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(filename)
    return filename

def cleanup_old_reports():
    current_time = datetime.now()
    for file in os.listdir():
        if file.startswith('report_') and file.endswith('.xlsx'):  # Изменил расширение на xlsx
            file_time = datetime.fromtimestamp(os.path.getctime(file))
            if (current_time - file_time).total_seconds() > 3600:  # Старше часа
                try:
                    os.remove(file)
                except:
                    pass 